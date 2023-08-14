import pandas as pd
from openpyxl import Workbook, load_workbook

class Easycell:
    def __init__(self, filepath=None):
        self.filepath = filepath
        self.workbook = None

        if self.filepath:
            self.load_workbook()

    def create_workbook(self):
        self.workbook = Workbook()

    def load_workbook(self):
        self.workbook = load_workbook(self.filepath)

    def save_workbook(self, filepath=None):
        if filepath:
            self.filepath = filepath
        if self.workbook:
            self.workbook.save(self.filepath)

    def create_sheet(self, sheet_name):
        if self.workbook:
            self.workbook.create_sheet(title=sheet_name)

    def get_sheet(self, sheet_name):
        if self.workbook and sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]

    def read_sheet(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            return pd.DataFrame(sheet.values)

    def write_dataframe(self, sheet_name, dataframe):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            for r_idx, row in enumerate(dataframe.values, start=1):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

    def filter_by(self, sheet_name, column_name, sort_type='asc'):
        sheet = self.get_sheet(sheet_name)
        if sheet and column_name in sheet[1]:
            data = pd.DataFrame(sheet.values)
            header_row = data.iloc[0]
            data = data[1:]
            column_index = header_row[header_row == column_name].index[0] + 1

            data_frame = pd.DataFrame(data, columns=header_row)
            filtered_data = data_frame.sort_values(by=column_name, ascending=(sort_type == 'asc'))
            
            return filtered_data


    def prevent_workbook(self):
        if self.workbook:
            self.workbook.security.lockStructure = True

    def lock_sheet(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            sheet.protection = Protection(locked=True)

    def unlock_sheet(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            sheet.protection = Protection(locked=False)

    def lock_headers(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            for cell in sheet[1]:
                cell.protection = Protection(locked=True)

    def expand_cells(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            sheet.sheet_view.showGridLines = False

    def conditional_style(self, sheet_name, condition):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if condition(cell.value):
                        cell.style = "Good"

    def set_conditional_style(self, sheet_name, style_name, bg_color):
        sheet = self.get_sheet(sheet_name)
        if sheet:
            new_style = NamedStyle(name=style_name)
            new_style.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            sheet.parent.styles.add_named_style(new_style)


    def query_to_sheet(self, sheet_name, query, connection):
            sheet = self.get_sheet(sheet_name)
            if sheet:
                result = pd.read_sql_query(query, connection)
                self.write_dataframe(sheet_name, result)